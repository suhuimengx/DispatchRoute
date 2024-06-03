import openpyxl
from points import GetRealTimeArray, RodeArray

file = openpyxl.load_workbook("./路径距离.xlsx")
for i in range(23):
    # 获取工作簿
    dataSheet = file.worksheets[i]
    dataSheet['K1'] = "百度路径"
    dataSheet['L1'] = "百度路径reverse"
    dataSheet['M1'] = "高德路径"
    dataSheet['N1'] = "高德路径reverse"
    # print(dataSheet['A' + str(1)].value)
    # 遍历每种路径
    for j in range(1, 24 - i):
        # j+2为所在行数；j+i为当前行目标点的索引
        # if dataSheet.cell(j + 2, 9).value == "直达":
        start = RodeArray[i]
        end = RodeArray[j+i]
        result = GetRealTimeArray(start, end, 30)
        distance, location_array, location_array_amap = [value for value in result.values()]
        # 以格式化字符串形式存储路径点
        location_array_str = ';'.join([','.join([str(item_float) for item_float in item]) for item in location_array])
        location_array_amap_str = ';'.join([','.join([str(item_float) for item_float in item]) for item in location_array_amap])
        location_array.reverse()
        location_array_amap.reverse()
        location_array_str_reverse = ';'.join([','.join([str(item_float) for item_float in item]) for item in location_array])
        location_array_amap_str_reverse = ';'.join([','.join([str(item_float) for item_float in item]) for item in location_array_amap])
        dataSheet['K'+str(j+2)] = location_array_str
        dataSheet['L'+str(j+2)] = location_array_str_reverse
        dataSheet['M'+str(j+2)] = location_array_amap_str
        dataSheet['N'+str(j+2)] = location_array_amap_str_reverse

file.save("./路径距离.xlsx")


'''
读取特定单元格数据
openpyxl.load_workbook(filename): 打开指定的Excel文件并加载工作簿。
workbook.cell(row, column): 获取特定单元格的值。
workbook.cell(row, column, value): 设置特定单元格的值。

往特定单元格增添数据
workbook.cell(row, column, value): 设置特定单元格的值。
workbook.cell(row, column, value, hidden=False): 设置特定单元格的值，并指定是否隐藏单元格。
workbook.cell(row, column, value, height=None, width=None): 设置特定单元格的值，并指定单元格的高度和宽度。

修改数据
workbook.cell(row, column, value): 设置特定单元格的值。
workbook.cell(row, column, value, hidden=False): 设置特定单元格的值，并指定是否隐藏单元格。
workbook.cell(row, column, value, height=None, width=None): 设置特定单元格的值，并指定单元格的高度和宽度。

下面是一个简单的示例，演示如何使用openpyxl来读取和修改Excel文件中的数据：
from openpyxl import load_workbook
# 打开工作簿
workbook = load_workbook(filename='example.xlsx')
# 获取特定单元格的值
value = workbook.cell(row=1, column=1).value
print(value)  # 输出：'A1'
# 设置特定单元格的值
workbook.cell(row=1, column=1, value='New Value')
# 保存工作簿
workbook.save(filename='example.xlsx')
注意：在处理Excel文件时，建议使用openpyxl的最新版本，以获取更多功能和更好的性能。
'''
